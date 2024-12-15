from openpyxl import load_workbook
from docx import Document
import fitz
import os
import tempfile
import xlrd
import textract
    
def parse_pdf(file_path):
    try:
        pdf_document = fitz.open(file_path)
        text = ''
        for page in pdf_document:
            text += page.get_text()
        return text
    except Exception as e:
        print(f'解析 {file_path} 失败: {e}')
        return ''
    
def parse_docx(file_path):
    try:
        doc = Document(file_path)
        text = "\n".join(para.text for para in doc.paragraphs)
        return text
    except Exception as e:
        return ''
    
def parse_doc(file_path):
    try:
        return parse_docx(file_path)
    except Exception:
        try:
            return textract.process(file_path).decode('utf-8')
        except Exception as e:
            print(f"解析 {file_path} 失败: {e}")
            return ''
    
def parse_xlsx(file_path):
    try:
        workbook = load_workbook(file_path, data_only=True)
        text_content = []
        
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                text_content.append(" ".join(str(cell) if cell is not None else '' for cell in row))
        
        return "\n".join(text_content)
    except Exception as e:
        print(f"解析 {file_path} 失败: {e}")
        return ''

def parse_xls(file_path):
    try:
        workbook = xlrd.open_workbook(file_path)
        text_content = []
        for sheet in workbook.sheets():
            for row_idx in range(sheet.nrows):
                row = sheet.row(row_idx)
                text_content.append(" ".join(str(cell.value) if cell.value != '' else '' for cell in row))
        return "\n".join(text_content)
    except Exception as e:
        print(f"解析 {file_path} 失败: {e}")
        return ''
    
def parse(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".doc":
        return parse_doc(file_path)
    elif ext == ".docx":
        return parse_docx(file_path)
    elif ext == ".xls":
        return parse_xls(file_path)
    elif ext == ".xlsx":
        return parse_xlsx(file_path)
    elif ext == ".pdf":
        return parse_pdf(file_path)
    else:
        print("不支持的文件格式")
        return ''
    
if __name__ == "__main__":
    doc_path = "example_files/example.doc"
    docx_path = "example_files/example.docx"
    xls_path = "example_files/example.xls"
    xlsx_path = "example_files/example.xlsx"
    pdf_path = "example_files/example.pdf"
    
    print(f'解析 {doc_path}:')
    print(parse_doc(doc_path))

    print(f'解析 {docx_path}:')
    print(parse_docx(docx_path))

    print(f'解析 {xls_path}:')
    print(parse_xls(xls_path))

    print(f'解析 {xlsx_path}:')
    print(parse_xlsx(xlsx_path))

    print(f'解析 {pdf_path}:')
    print(parse_pdf(pdf_path))